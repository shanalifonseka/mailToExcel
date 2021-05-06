from tkinter import *
from tkinter import ttk
import win32com.client
import os
from tkinter import messagebox
from tkcalendar import *
from datetime import *
from datetime import date
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Color, Font, PatternFill
from Classes import header
import re
import sqlite3
from tkinter import filedialog

#filename = filedialog.askopenfilename(initialdir = "/",title = "Select a File", filetypes = [("Excel files", ".xlsx .xls")])

CalendarStatus = 0
HeaderList = []

def GetHeaders():
    global HeaderList
    con = sqlite3.connect('FormExporter.db')
    cur = con.cursor()
    for row in cur.execute('SELECT nOrder_ID, vColumn_Name, vFormat_Component, vFormat_Component_After FROM Header ORDER BY nOrder_ID ASC'):
        HeaderList.append(header(row[0], row[1], row[2], row[3]))

def SaveTo(entry_text):
    filename = filedialog.askopenfilename(initialdir = "/",title = "Select a File", filetypes = [("Excel files", ".xlsx .xls")])
    #filename = filedialog.askopenfilename(initialdir = "/",title = "Select a File", filetypes = [("Excel files", ".xlsx .xls")])
    entry_text.set(filename)

def ChangeMenue(root, selection):
    if selection == 0:
        background_label.configure(image = photo_background_export_forms)
    else:
        background_label.configure(image = photo_background_edit_format)

def Cal(calender_root, entry_text, type, master = None):

    global CalendarStatus

    if CalendarStatus == 0:

        if type == 0:     
            lv_x = calender_root.winfo_rootx()+13
            lv_y = calender_root.winfo_rooty()+165
        else:
            lv_x = calender_root.winfo_rootx()+13
            lv_y = calender_root.winfo_rooty()+232

        calender_root = Toplevel(root)
        calender_window_height = 188
        calender_window_width = 250
        calender_root.geometry("{}x{}+{}+{}".format(calender_window_width, calender_window_height, lv_x, lv_y))
        calender_root.resizable(width=False, height=False)
        calender_root.overrideredirect(1)

        #photo_background_calendar = PhotoImage(file = "background-calendar.png")

        #background_calendar_label = Label(calender_root, image = photo_background_calendar)
        #background_calendar_label.place(x = 0, y = 0, relwidth = 1, relheight = 1)

        if str(entry_text.get()) == "":
            today = date.today()
            day = today.day
            month = today.month
            year = today.year
        else:
            entrydate = entry_text.get()
            day = int(datetime.strptime(entrydate, '%Y-%m-%d').strftime('%d'))
            month = int(datetime.strptime(entrydate, '%Y-%m-%d').strftime('%m'))
            year = int(datetime.strptime(entrydate, '%Y-%m-%d').strftime('%Y'))

        calendar_from_to = Calendar(calender_root, selectmode = "day", year = year, month = month, day = day, background = "#5e2e72", selectbackground = "#5e2e72",font = ("Calibri", 10) )
        calendar_from_to.place(x = 0, y = 0)
        calendar_from_to.bind("<<CalendarSelected>>", lambda e: CalendarSetDate(calender_root, calendar_from_to, entry_text))  

        #button_cancel =  Label(image = photo_background_export_forms_icon_mails_to, bg = "#de7878")
        #button_select.place(x=25, y=200)
        #button_cancel.bind("<Button>", lambda e: CalendarDestroy(calender_root))

        CalendarStatus = 1

def CalendarDestroy(calender_root, master = None):
    global CalendarStatus
    CalendarStatus = 0
    calender_root.destroy()

def CalendarSetDate(calender_root, calendar_from_to, entry_text, master = None):
    global CalendarStatus
    entry_text.set(datetime.strptime(calendar_from_to.get_date(), '%m/%d/%y').strftime('%Y-%m-%d'))
    CalendarStatus = 0
    calender_root.destroy()

def Popup():
    messagebox.showinfo("Sucessful", "Form{s) exported sucessfully")

#Check with valid conditions for urls in string
def FindURL(string):
    regex = r"(?i)\b((?:https?://|www\d{0,3}[.]|[a-z0-9.\-]+[.][a-z]{2,4}/)(?:[^\s()<>]+|\(([^\s()<>]+|(\([^\s()<>]+\)))*\))+(?:\(([^\s()<>]+|(\([^\s()<>]+\)))*\)|[^\s`!()\[\]{};:'\".,<>?«»“”‘’]))"
    url = re.findall(regex,string)      
    return [x[0] for x in url]

#Check with valid conditions for email in string
def FindEmail(string):
    regex = '^(\w|\.|\_|\-)+[@](\w|\_|\-|\.)+[.]\w{2,3}$'
    email = re.findall(regex,string)      
    return [x[0] for x in email]

def Export(entry_text):
    # Call a Workbook() function of openpyxl to create a new blank Workbook object 
    wb = openpyxl.Workbook() 

    # Get workbook active sheet from the active attribute 
    sheet = wb.active 

    #Read content from the .txt file which act as a database
    with open('Headers.txt') as f:
        headerContents = f.read()

    GetHeaders()
    for header in HeaderList:
        sheet.cell(row = 1, column = header.orderID).value = header.columnName

    outlook=win32com.client.Dispatch("Outlook.Application").GetNameSpace("MAPI")
    inbox=outlook.GetDefaultFolder(6) #Inbox default index value is 6
    message=inbox.Items
    message2=message.GetLast()
    subject=message2.Subject
    body=message2.body
    date=message2.senton.date()
    sender=message2.Sender
    attachments=message2.Attachments
    rowID = 2
    for m in message:
        print(m.senton.date())
        if m.Subject=="New Form Entry: Student Project Placement Form":
            #Provide location information for the cell and set value for detail cells despite removed column
            index = 1
            colIndex = index
            s = m.body
            for header in HeaderList:
                start = s.find(header.columnFormatComponent)+len(header.columnFormatComponent)
                end = s.find(header.columnFormatComponentAfter)

                valueInput = s[start:end].strip()
                sheet.cell(row = rowID, column = header.orderID).value = valueInput
                
                url = FindURL(valueInput)
                if len(url) == 1:
                        sheet.cell(row = rowID, column = header.orderID).value = '=HYPERLINK("{}", "{}")'.format(url[0], valueInput) #valueInput.partition("<")[0]
                        sheet.cell(row = rowID, column = header.orderID).style = 'Hyperlink'

            rowID = rowID + 1

    #Set column width to fit the text and aligment of the text to top and wrapped
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter# Get the column name
        for cell in col:
            if cell.coordinate in sheet.merged_cells: # not check merge_cells
                continue
            try: # Necessary to avoid error on empty cells
                if str("\n") not in str(cell.value):
                    sheet[str(cell.coordinate)].alignment = Alignment(vertical= "top")
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                else:
                    sheet[str(cell.coordinate)].alignment = Alignment(wrapText=True, vertical= "top")
                    x = str(cell.value).split("\n")
                    for xi in x:
                        if len(xi) > max_length:
                            max_length = len(xi)

            except:
                pass

        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    #Get path from the .txt and save .xls file to the location
    path = headerContents[headerContents.find("%0P%")+len("%0P%"):headerContents.find("%0PE%")].strip()
    wb.save(path)
    Popup()


root = Tk()
root.title('Form Exporter')
root.iconbitmap('icon_main.ico')
root.resizable(width=False, height=False)

window_height = 400
window_width = 275

screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

x_cordinate = int((screen_width/2) - (window_width/2))
y_cordinate = int((screen_height/2) - (window_height/2))

root.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))

photo_background_export_forms = PhotoImage(file = "background-export-forms.png")
photo_background_export_forms_icon = PhotoImage(file = "background-export-forms-icon.png")
photo_background_export_forms_icon_mails_to = PhotoImage(file = "background-export-forms-icon-mails-to.png")
photo_background_export_forms_icon_export = PhotoImage(file = "background-export-forms-icon-export.png")
photo_background_edit_format = PhotoImage(file = "background-edit-format.png")
photo_background_edit_format_icon = PhotoImage(file = "background-edit-format-icon.png")

background_label = Label(root, image = photo_background_export_forms)
background_label.place(x = 0, y = 0, relwidth = 1, relheight = 1)

save_to_entry_text = StringVar()
label_Export_Forms_save_to_entry = Label(textvariable = save_to_entry_text, bg = "#ffffff", font = ("Calibri", 10))
label_Export_Forms_save_to_entry.place(x=60, y=68)
label_Export_Forms_save_to_button = Label(image = photo_background_export_forms_icon_mails_to, bg = "#de7878")
label_Export_Forms_save_to_button.place(x=23, y=70)
label_Export_Forms_save_to_button.bind("<Button>", lambda e: SaveTo(save_to_entry_text))

mail_from_entry_text = StringVar()
label_Export_Forms_mail_from_entry = Label(textvariable = mail_from_entry_text, bg = "#ffffff", font = ("Calibri", 10))
label_Export_Forms_mail_from_entry.place(x=60, y=135)
label_Export_Forms_mail_from_button = Label(image = photo_background_export_forms_icon_mails_to, bg = "#de7878")
label_Export_Forms_mail_from_button.place(x=23, y=137)
label_Export_Forms_mail_from_button.bind("<Button>", lambda e: Cal(root, mail_from_entry_text, 0))

mail_to_entry_text = StringVar()
label_Export_Forms_mail_to_entry = Label(textvariable = mail_to_entry_text, bg = "#ffffff", font = ("Calibri", 10))
label_Export_Forms_mail_to_entry.place(x=60, y=202)
label_Export_Forms_mail_to_button = Label(image = photo_background_export_forms_icon_mails_to, bg = "#de7878")
label_Export_Forms_mail_to_button.place(x=23, y=204)
label_Export_Forms_mail_to_button.bind("<Button>", lambda e: Cal(root, mail_to_entry_text, 1))

label_Export_Forms_export_button = Label(image = photo_background_export_forms_icon_export, bg = "#ffffff")
label_Export_Forms_export_button.place(x=9, y=260)
label_Export_Forms_export_button.bind("<Button>", lambda e: Export(save_to_entry_text))

label_Edit_Format = Label(image = photo_background_edit_format_icon, bg = "#e6e6e6")
label_Edit_Format.place(x=165, y=355)
label_Edit_Format.bind("<Button>", lambda e: ChangeMenue(root, 1))

label_Export_Forms = Label(image = photo_background_export_forms_icon, bg = "#e6e6e6")
label_Export_Forms.place(x=68, y=355)
label_Export_Forms.bind("<Button>", lambda e: ChangeMenue(root, 0))

root.mainloop()