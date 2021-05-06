from tkinter import *
from tkinter import ttk
import win32com.client
import os
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Color, Font, PatternFill
import re
import numpy
from tkinter import messagebox
from tkcalendar import *
from datetime import *
from datetime import date

def Splash(splash_root):
    splash_root = Tk()
    splash_root.title('Form Exporter')
    splash_root.geometry("200x50")
    label_splash = Label(splash_root, text = "Export")
    splash_root.pack()

def SplashDestroy(splash_root):
    splash_root.destroy()

def Popup():
    messagebox.showinfo("Sucessful", "Form{s) exported sucessfully")

def Cal(calender_root, entry_text, master = None):      
    calender_root = Toplevel(root)
    calender_root.title('Set Date')
    calender_root.iconbitmap('icons/ico/icon_main.ico')
    calender_window_height = 235
    calender_window_width = 250
    calender_screen_width = calender_root.winfo_screenwidth()
    calender_screen_height = calender_root.winfo_screenheight()
    calender_x_cordinate = int((calender_screen_width/2) - (calender_window_width/2))
    calender_y_cordinate = int((calender_screen_height/2) - (calender_window_height/2))
    calender_root.geometry("{}x{}+{}+{}".format(calender_window_width, calender_window_height, calender_x_cordinate, calender_y_cordinate))
    calender_root.resizable(width=False, height=False)

    if str(entry_text.get()) == "":
        today = date.today()
        day = today.day
        month = today.month
        year = today.year
    else:
        entrydate = entry_text.get()
        day = int(datetime.strptime(entrydate, '%d/%m/%Y').strftime('%d'))
        month = int(datetime.strptime(entrydate, '%d/%m/%Y').strftime('%m'))
        year = int(datetime.strptime(entrydate, '%d/%m/%Y').strftime('%Y'))

    calendar_from_to = Calendar(calender_root, selectmode = "day", year = year, month = month, day = day)
    calendar_from_to.pack()

    frame_row_calender_button = Frame(calender_root)
    button_select = Button(frame_row_calender_button, text = "Select", compound=LEFT, height = 100, width = 100, bg = '#b3b3b3')
    button_select.bind("<Button>", lambda e: CalendarGetDate(calender_root, calendar_from_to, entry_text))
    button_select.pack(side = RIGHT, pady = 10)

    button_cancel = Button(frame_row_calender_button, text = "Cancel", compound=LEFT, height = 100, width = 100, bg = '#b3b3b3')
    button_cancel.pack(side = RIGHT, pady = 10)
    button_cancel.bind("<Button>", lambda e: CalendarDestroy(calender_root))
    frame_row_calender_button.pack()

def CalendarDestroy(calender_root, master = None):
    calender_root.destroy()

def CalendarGetDate(calender_root, calendar_from_to, entry_text, master = None):
    entry_text.set(calendar_from_to.get_date()) 
    calender_root.destroy()

#Check with valid conditions for urls in string
def FindURL(string):
    regex = r"(?i)\b((?:https?://|www\d{0,3}[.]|[a-z0-9.\-]+[.][a-z]{2,4}/)(?:[^\s()<>]+|\(([^\s()<>]+|(\([^\s()<>]+\)))*\))+(?:\(([^\s()<>]+|(\([^\s()<>]+\)))*\)|[^\s`!()\[\]{};:'\".,<>?«»“”‘’]))"
    url = re.findall(regex,string)      
    return [x[0] for x in url]

#Check with valid conditions for email in string
def FindEmail(string):
    regex = '^(\w|\.|\_|\-)+[@](\w|\_|\-|\.)+[.]\w{2,3}$'
    email = re.findall(regex,string)      
    return [x[0] for x in email]

def Export():
    # Call a Workbook() function of openpyxl to create a new blank Workbook object 
    wb = openpyxl.Workbook() 

    # Get workbook active sheet from the active attribute 
    sheet = wb.active 

    #Read content from the .txt file which act as a database
    with open('Headers.txt') as f:
        headerContents = f.read()

    #Assign number of column in form format
    headerLen = int(headerContents[headerContents.find("%0LL%")+len("%0LL%"):headerContents.find("%0LLE%")].strip())

    #Declare and assign form format columns to the array
    index = 1
    headerFormat = []
    while index <= headerLen:
        headerFormat.append(headerContents[headerContents.find("%"+str(index)+"F%")+len("%"+str(index)+"F%"):headerContents.find("%"+str(index)+"FE%")].strip())
        index = index + 1

    #Declare and assign form format columns display text to the array
    index = 1
    headerDisplay = []
    while index <= headerLen:
        headerDisplay.append(headerContents[headerContents.find("%"+str(index)+"D%")+len("%"+str(index)+"D%"):headerContents.find("%"+str(index)+"DE%")].strip())
        index = index + 1
    
    #Provide location information for the cell and set display text value for header cells despite removed column
    index = 1
    colIndex = index
    while index < headerLen:
        if headerDisplay[index-1] != "Removed":
            sheet.cell(row = 1, column = colIndex).value = headerDisplay[index-1]
            colIndex = colIndex + 1
        index = index + 1

    outlook=win32com.client.Dispatch("Outlook.Application").GetNameSpace("MAPI")
    inbox=outlook.GetDefaultFolder(6) #Inbox default index value is 6
    message=inbox.Items
    message2=message.GetLast()
    subject=message2.Subject
    body=message2.body
    date=message2.senton.date()
    sender=message2.Sender
    attachments=message2.Attachments
    rowID = 2
    for m in message:
        print(m.senton.date())
        if m.Subject=="New Form Entry: Student Project Placement Form":
            #Provide location information for the cell and set value for detail cells despite removed column
            index = 1
            colIndex = index
            s = m.body
            while index < headerLen:
                if headerDisplay[index-1] != "Removed":
                    start = headerFormat[index-1]
                    end = headerFormat[index]
                    valueInput = s[s.find(start)+len(start):s.find(end)].strip()
                    url = FindURL(valueInput)
                    sheet.cell(row = rowID, column = colIndex).value = valueInput
                    if len(url) == 1:
                        sheet.cell(row = rowID, column = colIndex).value = '=HYPERLINK("{}", "{}")'.format(url[0], valueInput) #valueInput.partition("<")[0]
                        sheet.cell(row = rowID, column = colIndex).style = 'Hyperlink'
                    colIndex = colIndex + 1
                index = index + 1

            rowID = rowID + 1

    #Set column width to fit the text and aligment of the text to top and wrapped
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter# Get the column name
        for cell in col:
            if cell.coordinate in sheet.merged_cells: # not check merge_cells
                continue
            try: # Necessary to avoid error on empty cells
                if str("\n") not in str(cell.value):
                    sheet[str(cell.coordinate)].alignment = Alignment(vertical= "top")
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                else:
                    sheet[str(cell.coordinate)].alignment = Alignment(wrapText=True, vertical= "top")
                    x = str(cell.value).split("\n")
                    for xi in x:
                        if len(xi) > max_length:
                            max_length = len(xi)

            except:
                pass

        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    #Get path from the .txt and save .xls file to the location
    path = headerContents[headerContents.find("%0P%")+len("%0P%"):headerContents.find("%0PE%")].strip()
    wb.save(path)
    Popup()
   

root = Tk()
root.title('Form Exporter')
root.iconbitmap('icon_main.ico')
root.resizable(width=False, height=False)

window_height = 400
window_width = 275

screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

x_cordinate = int((screen_width/2) - (window_width/2))
y_cordinate = int((screen_height/2) - (window_height/2))

root.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))

tab_Main = ttk.Notebook(root)
tab_Main.pack(pady = 15)

frame_export_forms = Frame(tab_Main, width = 300, height = 400)
frame_edit_format = Frame(tab_Main, width = 300, height = 400)

frame_export_forms.pack(fill = "both", expand = 1)
frame_edit_format.pack(fill = "both", expand = 1)

tab_Main.add(frame_export_forms, text = "Export Forms")
tab_Main.add(frame_edit_format, text = "Edit Format")

#photo_open = PhotoImage(file = "cil-calendar-check.png")

frame_row_path = Frame(frame_export_forms)
label_path = Label(frame_row_path, text = "Path: \t")
label_path.pack(side = LEFT, pady = 10)
entry_path = Entry(frame_row_path)
entry_path.pack(side = LEFT)
button_path = Button(frame_row_path, bg = '#b3b3b3')
button_path.pack(side = LEFT)
button_path.bind("<Button>", lambda e: Cal(root))
frame_row_path.pack()

#photo_cal = PhotoImage(file = "cil-calendar-check.png")

frame_row_from = Frame(frame_export_forms)
label_from = Label(frame_row_from, text = "From: \t")
label_from.pack(side = LEFT, pady = 10)
entry_from_text = StringVar()
entry_from = Entry(frame_row_from, textvariable = entry_from_text)
entry_from.pack(side = LEFT)
button_from = Button(frame_row_from, bg = '#b3b3b3')
button_from.pack(side = LEFT)
button_from.bind("<Button>", lambda e: Cal(root, entry_from_text))
frame_row_from.pack()

frame_row_to = Frame(frame_export_forms)
label_to = Label(frame_row_to, text = "To: \t")
label_to.pack(side = LEFT, pady = 10)
entry_to_text = StringVar()
entry_to = Entry(frame_row_to, textvariable = entry_to_text)
entry_to.pack(side = LEFT)
button_to = Button(frame_row_to, bg = '#b3b3b3')
button_to.pack(side = LEFT)
button_to.bind("<Button>", lambda e: Cal(root, entry_to_text))
frame_row_to.pack()

#photo_export = PhotoImage(file = "cli-export-mail.png")

frame_row_export = Frame(frame_export_forms)
button_export = Button(frame_row_export, text = " Export", compound=LEFT, height = 20, width = 100, command = Export, bg = '#b3b3b3')
button_export.pack(side = RIGHT, pady = 10)
frame_row_export.pack()

root.mainloop()